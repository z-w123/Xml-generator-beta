#!/usr/bin/python3
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import lxml.etree
import lxml.builder
import xlrd
from yattag import Doc, indent
import argparse, hashlib, os, subprocess, sys, time
from datetime import datetime
import csv
#03/01/2021 update: for loop to trim reads + assemblies -> v4 spreadsheet (without the 'submission tool field')
#TODO: confirm that this works with output of covid-excel-utils script

parser = argparse.ArgumentParser(prog='ena-metadata-xml-generator.py', formatter_class=argparse.RawDescriptionHelpFormatter,
                                     epilog="""
        + ============================================================ +
        |  European Nucleotide Archive (ENA) Analysis Submission Tool  |
        |                                                              |
        |  Tool to register study and sample metadata to an ENA project  |
        |  , mainly in the drag and drop tool context.                           |
        + =========================================================== +
        """)
parser.add_argument('-u', '--username', help='Webin submission account username (e.g. Webin-XXXXX)', type=str, required=True)
parser.add_argument('-p', '--password', help='password for Webin submission account', type=str, required=True)
parser.add_argument('-t', '--test', help='Specify whether to use ENA test server for submission', action='store_true')
parser.add_argument('-f', '--file', help='path for the metadata spreadsheet', type=str, required=True)
args = parser.parse_args()

import pandas as pd
import fnmatch #module for unix style pattern matching
import os #module in python provides functions for interacting with the operating system
import glob #module is used to retrieve files/pathnames matching a specified pattern
os.listdir(".") #list files and dirs in wd - make sure you are in the one where the user metadata spreadsheet will be found
files_xlsx = glob.glob(args.file) #should we accept other spreadsheet extensions?

for f in files_xlsx:
    if fnmatch.fnmatch(f, '*genome*'):
        print('you are using an assembly spreadsheet')
        df = pd.read_excel(f, usecols="L:AV", header=1, sheet_name='Sheet1') #col range suits v4 
        df = df.iloc[2:,]
        df.dropna(axis=0, how='all', inplace=True)
        df.insert(7,"submission_tool",'drag and drop uploader tool',allow_duplicates=True) #study #to inject constant into trimmed df
        df.insert(24,"submission_tool",'drag and drop uploader tool',allow_duplicates=True) #sample
        df.insert(26,"sample capture status",'active surveillance in response to outbreak',allow_duplicates=False)
        df.iloc[0, df.columns.get_loc('submission_tool')] = '' #effectively removes the 'drag and drop uploader tool' value from units row, by replacing with NaN
        df.iloc[0, df.columns.get_loc('sample capture status')] = '' #see above
        df.rename(columns = {'collecting institute':'collecting institution'}, inplace = True)
        df["release_date"] = pd.to_datetime(df["release_date"]).dt.strftime( "%Y-%m-%d")
        df["collection date"] = pd.to_datetime(df["collection date"]).dt.strftime("%Y-%m-%d")
        df["receipt date"] = pd.to_datetime(df["receipt date"]).dt.strftime("%Y-%m-%d")
        print('user sample & study data below:')
        print(df)
        df.to_excel(r'trimmed_assembly_study_sample_metadata_22_feb.xlsx', index=False)
    elif fnmatch.fnmatch(f, '*raw_reads*'):
        print('you are using a raw reads spreadsheet')
        df = pd.read_excel(f, usecols="B:AL", header=1, sheet_name='Sheet1') #col range suits v4
        df = df.iloc[2:,]
        df.dropna(axis=0, how='all', inplace=True)
        df.insert(7,"submission_tool",'drag and drop uploader tool',allow_duplicates=True) #study #to inject constant into trimmed df
        df.insert(24,"submission_tool",'drag and drop uploader tool',allow_duplicates=True) #sample
        df.insert(26,"sample capture status",'active surveillance in response to outbreak',allow_duplicates=False)
        df.iloc[0, df.columns.get_loc('submission_tool')] = '' #effectively removes the 'drag and drop uploader tool' value from units row, by replacing with NaN
        df.iloc[0, df.columns.get_loc('sample capture status')] = '' #see above
        df.rename(columns = {'collecting institute':'collecting institution'}, inplace = True) #####temp fix for collecting institute error
        df["release_date"] = pd.to_datetime(df["release_date"]).dt.strftime( "%Y-%m-%d")
        df["collection date"] = pd.to_datetime(df["collection date"]).dt.strftime("%Y-%m-%d")
        df["receipt date"] = pd.to_datetime(df["receipt date"]).dt.strftime("%Y-%m-%d")
        print('user sample & study data below:')
        print(df)
        df.to_excel(r'trimmed_raw_reads_study_sample_metadata_22_feb.xlsx', index=False)
    else:
        print('you have used an unsupported spreadsheet, please try again')


wb = load_workbook("trimmed_raw_reads_study_sample_metadata_22_feb.xlsx")
ws = wb.worksheets[0]
doc, tag, text = Doc().tagtext()
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
doc.asis(xml_header)

if ws['B3'].value !=None:
    with tag('STUDY_SET'):
        for row_study in ws.iter_rows(min_row=3, min_col=2, max_col=8, values_only=True):
            found_study = False
            for y in row_study:
                if y != None:
                    found_study = True
            if found_study == True:
                first_study = row_study[0:6]
                all_study = row_study[6:]
                with tag('STUDY', alias=first_study[0]):
                    with tag("DESCRIPTOR"):
                        with tag("STUDY_TITLE"):
                            text(first_study[3])
                        doc.stag('STUDY_TYPE', existing_study_type="Other")
                        with tag("STUDY_ABSTRACT"):
                            text(first_study[5])
                        with tag("CENTER_PROJECT_NAME"):
                            text(first_study[4])
                    with tag('STUDY_ATTRIBUTES'):
                        for z in range(len(all_study)):
                            y = all_study[z];
                            if y != None:
                                with tag("STUDY_ATTRIBUTE"):
                                    with tag("TAG"):
                                        text(ws[1][z+2+len(first_study)].value)
                                    with tag("VALUE"):
                                        text(str(y))


    result_study = indent(
        doc.getvalue(),
        #indentation = '    ',
        indent_text = False
    )

    with open("study.xml", "w") as f:
        f.write(result_study)

# Creating sample xml
# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'

if ws['K3'].value !=None:
    doc.asis(xml_header)
    with tag('SAMPLE_SET'):
        for row in ws.iter_rows(min_row=3, min_col=11, values_only=True):
            found = False
            for x in row:
                if x != None:
                    found = True
            if found == True:
                first = row[0:5]
                all = row[5:]
                with tag('SAMPLE', alias=first[0]):
                    with tag("TITLE"):
                        text(first[3])
                    with tag('SAMPLE_NAME'):
                        with tag("TAXON_ID"):
                            text(first[1])
                        with tag("SCIENTIFIC_NAME"):
                            text(first[2])
                    with tag("DESCRIPTION"):
                        text(first[4])
                    with tag('SAMPLE_ATTRIBUTES'):
                        for y in range(len(all)):
                            x = all[y];
                            if x != None:
                                with tag("SAMPLE_ATTRIBUTE"):
                                    with tag("TAG"):
                                        text(ws[1][y+10+len(first)].value)
                                    with tag("VALUE"):
                                        text(str(x))
                                    if ws[2][y+10+len(first)].value != None:
                                        with tag("UNITS"):
                                            text(ws[2][y+10+len(first)].value)
                        with tag("SAMPLE_ATTRIBUTE"):
                            with tag("TAG"):
                                text("ENA-CHECKLIST")
                            with tag("VALUE"):
                                text("ERC000033")




    result = indent(
        doc.getvalue(),
        #indentation = '    ',
        indent_text = False
    )

    with open("sample.xml", "w") as f:
        f.write(result)

#creating the submission xml

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
#xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'

doc.asis(xml_header)
#doc.asis(xml_schema)

with tag('SUBMISSION_SET'):
    with tag('SUBMISSION'):
        with tag("ACTIONS"):
            with tag('ACTION'):
                doc.stag('ADD')
            if ws['I3'].value != None:
                with tag('ACTION'):
                    doc.stag('HOLD', HoldUntilDate=str(ws['I3'].value))

result_s = indent(
    doc.getvalue(),
    indentation='    ',
    indent_text=True
)

with open("submission.xml", "w") as f:
    f.write(result_s)

# submission command
if ws['B3'].value == None:
    if args.test is True:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "SAMPLE=@sample.xml"  "https://wwwdev.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(
            args.username, args.password)

    if args.test is False:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "SAMPLE=@sample.xml"  "https://www.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(
            args.username, args.password)
elif ws['K3'].value == None:
    if args.test is True:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "STUDY=@study.xml" "https://wwwdev.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(args.username, args.password)

    if args.test is False:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "STUDY=@study.xml" "https://www.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(args.username, args.password)
else:
    if args.test is True:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "SAMPLE=@sample.xml" -F "STUDY=@study.xml" "https://wwwdev.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(args.username, args.password)

    if args.test is False:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "SAMPLE=@sample.xml" -F "STUDY=@study.xml" "https://www.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(args.username, args.password)

sp = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
out, err = sp.communicate()

print("-" * 100)
print("CURL submission command: \n")
print(command)
print("Returned output: \n")
print(out.decode())
print("-" * 100)