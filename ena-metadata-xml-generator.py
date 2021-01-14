from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import lxml.etree
import lxml.builder
import xlrd
from yattag import Doc, indent
import argparse, hashlib, os, subprocess, sys, time
from datetime import datetime


parser = argparse.ArgumentParser(prog='ena-metadata-xml-generator.py', formatter_class=argparse.RawDescriptionHelpFormatter,
                                     epilog="""
        + ============================================================ +
        |  European Nucleotide Archive (ENA) Analysis Submission Tool  |
        |                                                              |
        |  Tool to register study and sample metadata to an ENA project  |
        |  , mainly in the drag and drop tool context.                           |
        + =========================================================== +
        """)
parser.add_argument('-u', '--username', help='Webin submission account username (e.g. Webin-XXXXX)', type=str, required=True)
parser.add_argument('-p', '--password', help='password for Webin submission account', type=str, required=True)
parser.add_argument('-t', '--test', help='Specify whether to use ENA test server for submission', action='store_true')
args = parser.parse_args()


wb = load_workbook("uploader_tool_metadata_v3_raw_reads_notes_only.xlsx")
ws = wb.worksheets[0]

#Creating the study xml
# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
doc.asis(xml_header)
if ws['C6'].value !=None:
    with tag('STUDY_SET'):
        for row_study in ws.iter_rows(min_row=6, min_col=3, max_col=9, values_only=True):
            found_study = False
            for y in row_study:
                if y != None:
                    found_study = True
            if found_study == True:
                first_study = row_study[0:6]
                all_study = row_study[6:]
                with tag('STUDY', alias=first_study[0]):
                    with tag("DESCRIPTOR"):
                        with tag("STUDY_TITLE"):
                            text(first_study[3])
                        doc.stag('STUDY_TYPE', existing_study_type="Other")
                        with tag("STUDY_ABSTRACT"):
                            text(first_study[5])
                        with tag("CENTER_PROJECT_NAME"):
                            text(first_study[4])
                    with tag('STUDY_ATTRIBUTES'):
                        for y in all_study:
                            if y != None:
                                with tag("STUDY_ATTRIBUTE"):
                                    with tag("TAG"):
                                        text(ws[2][row_study.index(y)+2].value)
                                    with tag("VALUE"):
                                        text(str(y))


    result_study = indent(
        doc.getvalue(),
        #indentation = '    ',
        indent_text = False
    )

    with open("study.xml", "w") as f:
        f.write(result_study)

# Creating sample xml
# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'

if ws['L6'].value !=None:
    doc.asis(xml_header)
    with tag('SAMPLE_SET'):
        for row in ws.iter_rows(min_row=6, min_col=12, max_col=41, values_only=True):
            found = False
            for x in row:
                if x != None:
                    found = True
            if found == True:
                first = row[0:5]
                all = row[5:]
                with tag('SAMPLE', alias=first[0]):
                    with tag("TITLE"):
                        text(first[3])
                    with tag('SAMPLE_NAME'):
                        with tag("TAXON_ID"):
                            text(first[1])
                        with tag("SCIENTIFIC_NAME"):
                            text(first[2])
                    with tag("DESCRIPTION"):
                        text(first[4])
                    with tag('SAMPLE_ATTRIBUTES'):
                        for x in all:
                            if x != None:
                                with tag("SAMPLE_ATTRIBUTE"):
                                    with tag("TAG"):
                                        text(ws[2][row.index(x)+11].value)
                                    with tag("VALUE"):
                                        text(str(x))
                                    if ws[5][row.index(x)+11].value != None:
                                        with tag("UNITS"):
                                            text(ws[5][row.index(x)+11].value)
                        with tag("SAMPLE_ATTRIBUTE"):
                            with tag("TAG"):
                                text("ENA-CHECKLIST")
                            with tag("VALUE"):
                                text("ERC000033")




    result = indent(
        doc.getvalue(),
        #indentation = '    ',
        indent_text = False
    )

    with open("sample.xml", "w") as f:
        f.write(result)

#creating the submission xml

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
#xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'

doc.asis(xml_header)
#doc.asis(xml_schema)

with tag('SUBMISSION_SET'):
    with tag('SUBMISSION'):
        with tag("ACTIONS"):
            with tag('ACTION'):
                doc.stag('ADD')
            if ws['J6'].value != None:
                with tag('ACTION'):
                    doc.stag('HOLD', HoldUntilDate=str(ws['J6'].value))

result_s = indent(
    doc.getvalue(),
    indentation='    ',
    indent_text=True
)

with open("submission.xml", "w") as f:
    f.write(result_s)

# submission command
if ws['C6'].value == None:
    if args.test is True:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "SAMPLE=@sample.xml"  "https://wwwdev.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(
            args.username, args.password)

    if args.test is False:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "SAMPLE=@sample.xml"  "https://www.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(
            args.username, args.password)
elif ws['L6'].value == None:
    if args.test is True:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "STUDY=@study.xml" "https://wwwdev.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(args.username, args.password)

    if args.test is False:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "STUDY=@study.xml" "https://www.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(args.username, args.password)
else:
    if args.test is True:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "SAMPLE=@sample.xml" -F "STUDY=@study.xml" "https://wwwdev.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(args.username, args.password)

    if args.test is False:
        command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "SAMPLE=@sample.xml" -F "STUDY=@study.xml" "https://www.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(args.username, args.password)

sp = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
out, err = sp.communicate()

print("-" * 100)
print("CURL submission command: \n")
print(command)
print("Returned output: \n")
print(out.decode())
print("-" * 100)